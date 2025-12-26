from typing import Annotated, Optional, List
from fastapi import APIRouter, Depends, HTTPException, Query, File, Form, UploadFile
from fastapi.responses import StreamingResponse, FileResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func, or_, cast
from sqlalchemy.dialects.postgresql import JSONB
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from io import BytesIO
from datetime import datetime, date
from app.core.db import get_db
from app.core.permissions import PERM_STUDENTS_VIEW, PERM_STUDENTS_EDIT, PERM_ATTENDANCE_VIEW
from app.models.domain import Student
from app.models.finance import Transaction
from app.models.attendance import Attendance, GateLog, Session
from app.models.enums import StudentStatus, ContractStatus
from app.schemas.student import StudentRead, StudentCreate, StudentUpdate, StudentDebtInfo, StudentFullInfo, ParentRead
from app.schemas.contract import ContractRead
from app.schemas.transaction import TransactionRead
from app.schemas.attendance import AttendanceRead, GateLogRead
from app.schemas.common import DataResponse, PaginationMeta
from app.schemas.student_with_contract import StudentWithContractCreate, StudentWithContractResponse
from app.deps import require_permission, CurrentUser
from app.models.auth import User
from app.core.s3 import upload_image_to_s3, upload_pdf_to_s3
from app.utils.contract_pdf import ContractPDFGenerator

router = APIRouter(prefix="/students", tags=["Students"])


@router.get("/search", response_model=DataResponse[list[StudentRead]], dependencies=[Depends(require_permission(PERM_STUDENTS_VIEW))])
async def search_students(
    db: Annotated[AsyncSession, Depends(get_db)],
    query: str = Query(..., description="Search by first name, last name, contract number, phone, or parent email"),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
):
    """
    Comprehensive search for students by:
    - First name
    - Last name
    - Contract number
    - Phone number
    - Parent email
    """
    from app.models.domain import Contract, Parent

    # Search students by name or phone
    students_query = select(Student).where(
        or_(
            Student.first_name.ilike(f"%{query}%"),
            Student.last_name.ilike(f"%{query}%"),
            Student.phone.ilike(f"%{query}%"),
        )
    ).distinct()

    # Search by contract number
    contracts_result = await db.execute(
        select(Contract.student_id).where(Contract.contract_number.ilike(f"%{query}%"))
    )
    student_ids_from_contracts = [row[0] for row in contracts_result.fetchall()]

    # Search by parent email
    parents_result = await db.execute(
        select(Parent.student_id).where(Parent.email.ilike(f"%{query}%"))
    )
    student_ids_from_parents = [row[0] for row in parents_result.fetchall()]

    # Combine all student IDs
    all_student_ids = set(student_ids_from_contracts + student_ids_from_parents)

    # If we found students via contracts or parents, add them to the query
    if all_student_ids:
        students_query = select(Student).where(
            or_(
                Student.first_name.ilike(f"%{query}%"),
                Student.last_name.ilike(f"%{query}%"),
                Student.phone.ilike(f"%{query}%"),
                Student.id.in_(all_student_ids)
            )
        ).distinct()

    # Apply pagination
    offset = (page - 1) * page_size
    result = await db.execute(students_query.offset(offset).limit(page_size))
    students = result.scalars().all()

    # Count total results
    count_query = select(func.count(Student.id.distinct())).where(
        or_(
            Student.first_name.ilike(f"%{query}%"),
            Student.last_name.ilike(f"%{query}%"),
            Student.phone.ilike(f"%{query}%"),
            Student.id.in_(all_student_ids) if all_student_ids else False
        )
    )
    count_result = await db.execute(count_query)
    total = count_result.scalar() or 0

    return DataResponse(
        data=[StudentRead.model_validate(s) for s in students],
        meta=PaginationMeta(
            page=page,
            page_size=page_size,
            total=total,
            total_pages=(total + page_size - 1) // page_size if total > 0 else 0,
        ),
    )


@router.get("", response_model=DataResponse[list[StudentRead]], dependencies=[Depends(require_permission(PERM_STUDENTS_VIEW))])
async def get_students(
    db: Annotated[AsyncSession, Depends(get_db)],
    search: Optional[str] = None,
    group_id: Optional[int] = None,
    status: Optional[str] = None,
    archive_year: int | None = Query(None, description="Filter by archive year (defaults to current year)"),
    include_archived: bool = Query(False, description="Include archived students (default: only non-ARCHIVED)"),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
):
    """
    Get all students with optional filters.

    Default behavior:
    - Shows current year's students only
    - Excludes ARCHIVED students unless include_archived=true
    """
    from datetime import datetime as dt
    if archive_year is None:
        archive_year = dt.now().year

    query = select(Student).where(Student.archive_year == archive_year)

    # Default: exclude ARCHIVED students
    if not include_archived:
        query = query.where(Student.status != StudentStatus.ARCHIVED)

    if search:
        query = query.where(
            or_(
                Student.first_name.ilike(f"%{search}%"),
                Student.last_name.ilike(f"%{search}%"),
            )
        )
    if group_id:
        query = query.where(Student.group_id == group_id)
    if status:
        query = query.where(Student.status == status)

    offset = (page - 1) * page_size
    result = await db.execute(query.offset(offset).limit(page_size))
    students = result.scalars().all()

    count_query = select(func.count(Student.id)).where(Student.archive_year == archive_year)
    if not include_archived:
        count_query = count_query.where(Student.status != StudentStatus.ARCHIVED)
    if search:
        count_query = count_query.where(
            or_(
                Student.first_name.ilike(f"%{search}%"),
                Student.last_name.ilike(f"%{search}%"),
            )
        )
    if group_id:
        count_query = count_query.where(Student.group_id == group_id)
    if status:
        count_query = count_query.where(Student.status == status)

    count_result = await db.execute(count_query)
    total = count_result.scalar()

    return DataResponse(
        data=[StudentRead.model_validate(s) for s in students],
        meta=PaginationMeta(
            page=page,
            page_size=page_size,
            total=total,
            total_pages=(total + page_size - 1) // page_size,
        ),
    )


@router.get("/unpaid", response_model=DataResponse[list[StudentDebtInfo]], dependencies=[Depends(require_permission(PERM_STUDENTS_VIEW))])
async def get_unpaid_students(
    db: Annotated[AsyncSession, Depends(get_db)],
    year: Optional[int] = None,
    month: Optional[int] = Query(None, ge=1, le=12),
    months: Optional[str] = Query(None, description="Comma-separated months (e.g., '1,2,3' for Jan, Feb, Mar)"),
    from_date: Optional[date] = Query(None, description="Start date for date range filtering (YYYY-MM-DD)"),
    to_date: Optional[date] = Query(None, description="End date for date range filtering (YYYY-MM-DD)"),
    group_id: Optional[int] = None,
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
):
    """
    Get students who haven't paid for specified periods.

    Filters (use either year/month OR from_date/to_date):
    - year: Target year (defaults to current year)
    - month: Single month to check (1-12)
    - months: Multiple months as comma-separated string (e.g., "1,2,3")
    - from_date: Start date for date range (e.g., "2025-01-01")
    - to_date: End date for date range (e.g., "2025-03-31")
    - group_id: Filter by specific group

    Examples:
    - /unpaid?year=2025&month=1 - Debtors for January 2025
    - /unpaid?year=2025&months=1,2,3 - Debtors for Jan, Feb, Mar 2025
    - /unpaid?from_date=2025-01-01&to_date=2025-03-31 - Debtors for Q1 2025
    - /unpaid?year=2025 - Debtors for any month in 2025
    - /unpaid?year=2025&group_id=5 - Debtors in group 5 for 2025
    """
    from app.models.domain import Contract
    from app.models.enums import ContractStatus, PaymentStatus
    from datetime import date as date_type
    from dateutil.relativedelta import relativedelta

    # Default to current year
    today = date_type.today()

    # Determine which filtering mode to use: date range or year/month
    use_date_range = from_date is not None or to_date is not None

    # Parse target months as (year, month) tuples
    target_months = []

    if use_date_range:
        # Date range mode
        if from_date is None:
            from_date = date_type(today.year, 1, 1)  # Default to start of current year
        if to_date is None:
            to_date = today  # Default to today

        if from_date > to_date:
            raise HTTPException(status_code=400, detail="from_date must be before or equal to to_date")

        # Calculate all months in the date range
        current_date = from_date.replace(day=1)
        end_date = to_date.replace(day=1)

        while current_date <= end_date:
            target_months.append((current_date.year, current_date.month))
            current_date += relativedelta(months=1)
    else:
        # Year/month mode (existing logic)
        target_year = year if year is not None else today.year

        month_list = []
        if months:
            # Parse comma-separated months
            try:
                month_list = [int(m.strip()) for m in months.split(',') if m.strip()]
                # Validate months
                for m in month_list:
                    if m < 1 or m > 12:
                        raise HTTPException(status_code=400, detail=f"Invalid month: {m}. Must be between 1 and 12")
            except ValueError:
                raise HTTPException(status_code=400, detail="Invalid months format. Use comma-separated numbers (e.g., '1,2,3')")
        elif month is not None:
            # Single month specified
            month_list = [month]
        else:
            # No months specified - check all 12 months of the year
            month_list = list(range(1, 13))

        # Convert to (year, month) tuples
        target_months = [(target_year, m) for m in month_list]

    # Build student query with filters
    students_query = select(Student).where(Student.status == "active")

    if group_id:
        students_query = students_query.where(Student.group_id == group_id)

    students_result = await db.execute(students_query)
    students = students_result.scalars().all()

    debt_info_list = []
    for student in students:
        # Get all contracts for this student (including terminated ones)
        contracts_result = await db.execute(
            select(Contract).where(Contract.student_id == student.id)
        )
        contracts = contracts_result.scalars().all()

        if not contracts:
            continue

        # Calculate debt across all target months
        total_expected = 0
        total_paid = 0
        active_contracts_count = 0

        for target_year_month, target_month_num in target_months:
            # Calculate expected payment for this month
            month_expected = 0
            target_date = date_type(target_year_month, target_month_num, 1)

            for contract in contracts:
                # Determine the effective end date (earliest of end_date or terminated_at)
                effective_end_date = contract.end_date
                if contract.terminated_at:
                    termination_date = contract.terminated_at.date()
                    if termination_date < effective_end_date:
                        effective_end_date = termination_date

                # Check if contract was active during this target month
                # Contract start month
                contract_start_month = contract.start_date.replace(day=1)
                contract_end_month = effective_end_date.replace(day=1)

                # Check if target month falls within contract period
                if contract_start_month <= target_date <= contract_end_month:
                    month_expected += float(contract.monthly_fee)
                    # Count as active if it was active in any of the target months
                    if contract.status == ContractStatus.ACTIVE or (
                        contract.status == ContractStatus.TERMINATED and
                        contract.terminated_at and
                        contract.terminated_at.date() >= target_date
                    ):
                        active_contracts_count = max(active_contracts_count, 1)

            # Add to total expected
            total_expected += month_expected

            # Check if student has paid for this specific month
            # Use PostgreSQL @> operator to check if JSON array contains the target month
            # Cast payment_months to JSONB to avoid type mismatch
            transactions_result = await db.execute(
                select(func.sum(Transaction.amount)).where(
                    Transaction.student_id == student.id,
                    Transaction.status == PaymentStatus.SUCCESS,
                    Transaction.payment_year == target_year_month,
                    cast(Transaction.payment_months, JSONB).op('@>')(cast([target_month_num], JSONB))
                )
            )
            month_paid = transactions_result.scalar() or 0
            total_paid += float(month_paid)

        # If no expected payment for any month, skip
        if total_expected == 0:
            continue

        # Calculate total debt
        debt_amount = total_expected - total_paid

        if debt_amount > 0.01:  # Only include if debt > 1 cent (avoid floating point errors)
            debt_info_list.append(StudentDebtInfo(
                student=StudentRead.model_validate(student),
                total_expected=total_expected,
                total_paid=total_paid,
                debt_amount=debt_amount,
                active_contracts_count=len([c for c in contracts if c.status == ContractStatus.ACTIVE])
            ))

    # Sort by debt amount (highest first)
    debt_info_list.sort(key=lambda x: x.debt_amount, reverse=True)

    # Apply pagination
    offset = (page - 1) * page_size
    paginated_list = debt_info_list[offset:offset + page_size]
    total = len(debt_info_list)

    return DataResponse(
        data=paginated_list,
        meta=PaginationMeta(
            page=page,
            page_size=page_size,
            total=total,
            total_pages=(total + page_size - 1) // page_size,
        ),
    )


@router.get("/unpaid/export", dependencies=[Depends(require_permission(PERM_STUDENTS_VIEW))])
async def export_unpaid_students(
    db: Annotated[AsyncSession, Depends(get_db)],
    year: Optional[int] = None,
    month: Optional[int] = Query(None, ge=1, le=12),
    months: Optional[str] = Query(None, description="Comma-separated months (e.g., '1,2,3')"),
    from_date: Optional[date] = Query(None, description="Start date for date range filtering (YYYY-MM-DD)"),
    to_date: Optional[date] = Query(None, description="End date for date range filtering (YYYY-MM-DD)"),
    group_id: Optional[int] = None,
):
    """
    Export unpaid students data to Excel file with statistics.

    Same filters as /unpaid endpoint (use either year/month OR from_date/to_date):
    - year: Target year (defaults to current year)
    - month: Single month to check (1-12)
    - months: Multiple months as comma-separated string
    - from_date: Start date for date range (e.g., "2025-01-01")
    - to_date: End date for date range (e.g., "2025-03-31")
    - group_id: Filter by specific group
    """
    from app.models.domain import Contract, Group
    from app.models.enums import ContractStatus, PaymentStatus
    from datetime import date as date_type
    from dateutil.relativedelta import relativedelta

    # Default to current year
    today = date_type.today()

    # Determine which filtering mode to use: date range or year/month
    use_date_range = from_date is not None or to_date is not None

    # Parse target months as (year, month) tuples
    target_months = []

    if use_date_range:
        # Date range mode
        if from_date is None:
            from_date = date_type(today.year, 1, 1)
        if to_date is None:
            to_date = today

        if from_date > to_date:
            raise HTTPException(status_code=400, detail="from_date must be before or equal to to_date")

        # Calculate all months in the date range
        current_date = from_date.replace(day=1)
        end_date = to_date.replace(day=1)

        while current_date <= end_date:
            target_months.append((current_date.year, current_date.month))
            current_date += relativedelta(months=1)
    else:
        # Year/month mode
        target_year = year if year is not None else today.year

        month_list = []
        if months:
            try:
                month_list = [int(m.strip()) for m in months.split(',') if m.strip()]
                for m in month_list:
                    if m < 1 or m > 12:
                        raise HTTPException(status_code=400, detail=f"Invalid month: {m}. Must be between 1 and 12")
            except ValueError:
                raise HTTPException(status_code=400, detail="Invalid months format. Use comma-separated numbers")
        elif month is not None:
            month_list = [month]
        else:
            month_list = list(range(1, 13))

        target_months = [(target_year, m) for m in month_list]

    # Build student query with filters
    students_query = select(Student).where(Student.status == "active")

    if group_id:
        students_query = students_query.where(Student.group_id == group_id)

    students_result = await db.execute(students_query)
    students = students_result.scalars().all()

    # Get group name for filename if filtering by group
    group_name = ""
    if group_id:
        group_result = await db.execute(select(Group).where(Group.id == group_id))
        group = group_result.scalar_one_or_none()
        if group:
            group_name = f"_{group.name.replace(' ', '_')}"

    debt_info_list = []
    for student in students:
        contracts_result = await db.execute(
            select(Contract).where(Contract.student_id == student.id)
        )
        contracts = contracts_result.scalars().all()

        if not contracts:
            continue

        total_expected = 0
        total_paid = 0

        for target_year_month, target_month_num in target_months:
            month_expected = 0
            target_date = date_type(target_year_month, target_month_num, 1)

            for contract in contracts:
                effective_end_date = contract.end_date
                if contract.terminated_at:
                    termination_date = contract.terminated_at.date()
                    if termination_date < effective_end_date:
                        effective_end_date = termination_date

                contract_start_month = contract.start_date.replace(day=1)
                contract_end_month = effective_end_date.replace(day=1)

                if contract_start_month <= target_date <= contract_end_month:
                    month_expected += float(contract.monthly_fee)

            total_expected += month_expected

            # Cast payment_months to JSONB to avoid type mismatch
            transactions_result = await db.execute(
                select(func.sum(Transaction.amount)).where(
                    Transaction.student_id == student.id,
                    Transaction.status == PaymentStatus.SUCCESS,
                    Transaction.payment_year == target_year_month,
                    cast(Transaction.payment_months, JSONB).op('@>')(cast([target_month_num], JSONB))
                )
            )
            month_paid = transactions_result.scalar() or 0
            total_paid += float(month_paid)

        if total_expected == 0:
            continue

        debt_amount = total_expected - total_paid

        if debt_amount > 0.01:
            # Get group name for this student
            student_group_name = ""
            if student.group_id:
                student_group_result = await db.execute(select(Group).where(Group.id == student.group_id))
                student_group = student_group_result.scalar_one_or_none()
                if student_group:
                    student_group_name = student_group.name

            debt_info_list.append({
                "student_id": student.id,
                "first_name": student.first_name,
                "last_name": student.last_name,
                "phone": student.phone or "",
                "group": student_group_name,
                "total_expected": total_expected,
                "total_paid": total_paid,
                "debt_amount": debt_amount,
                "active_contracts": len([c for c in contracts if c.status == ContractStatus.ACTIVE])
            })

    # Sort by debt amount (highest first)
    debt_info_list.sort(key=lambda x: x["debt_amount"], reverse=True)

    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Unpaid Students"

    # Define header style
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    header_alignment = Alignment(horizontal="center", vertical="center")

    # Define headers
    headers = [
        "ID",
        "First Name",
        "Last Name",
        "Phone",
        "Group",
        "Expected Amount",
        "Paid Amount",
        "Debt Amount",
        "Active Contracts"
    ]

    # Write headers
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    # Write data
    for row_num, debt_info in enumerate(debt_info_list, 2):
        ws.cell(row=row_num, column=1, value=debt_info["student_id"])
        ws.cell(row=row_num, column=2, value=debt_info["first_name"])
        ws.cell(row=row_num, column=3, value=debt_info["last_name"])
        ws.cell(row=row_num, column=4, value=debt_info["phone"])
        ws.cell(row=row_num, column=5, value=debt_info["group"])
        ws.cell(row=row_num, column=6, value=debt_info["total_expected"])
        ws.cell(row=row_num, column=7, value=debt_info["total_paid"])
        ws.cell(row=row_num, column=8, value=debt_info["debt_amount"])
        ws.cell(row=row_num, column=9, value=debt_info["active_contracts"])

    # Adjust column widths
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 15
    ws.column_dimensions['I'].width = 18

    # Add summary row
    if debt_info_list:
        summary_row = len(debt_info_list) + 3
        ws.cell(row=summary_row, column=1, value="TOTAL").font = Font(bold=True)
        ws.cell(row=summary_row, column=6, value=sum(d["total_expected"] for d in debt_info_list)).font = Font(bold=True)
        ws.cell(row=summary_row, column=7, value=sum(d["total_paid"] for d in debt_info_list)).font = Font(bold=True)
        ws.cell(row=summary_row, column=8, value=sum(d["debt_amount"] for d in debt_info_list)).font = Font(bold=True)

    # Save to BytesIO
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)

    # Generate filename based on filter type
    if use_date_range:
        date_str = f"{from_date.strftime('%Y%m%d')}_{to_date.strftime('%Y%m%d')}"
        filename = f"unpaid_students_{date_str}{group_name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    else:
        # Extract just the month numbers for display
        month_nums = [m[1] for m in target_months]
        months_str = ",".join(map(str, month_nums)) if len(month_nums) <= 3 else "all"
        filename = f"unpaid_students_{target_year}_{months_str}{group_name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    # Return as streaming response
    return StreamingResponse(
        excel_file,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/comprehensive-export", dependencies=[Depends(require_permission(PERM_STUDENTS_VIEW))])
async def export_comprehensive_student_data(
    db: Annotated[AsyncSession, Depends(get_db)],
    from_date: Optional[date] = Query(None, description="Start date for payment history (YYYY-MM-DD)"),
    to_date: Optional[date] = Query(None, description="End date for payment history (YYYY-MM-DD)"),
    group_id: Optional[int] = Query(None, description="Filter by specific group"),
    status: Optional[str] = Query(None, description="Filter by student status (active, archived, etc.)"),
):
    """
    Export comprehensive student data to Excel file including:
    - Student information (name, phone, address, date of birth, status)
    - Contract details (number, start/end dates, monthly fee, status, termination info)
    - Payment history (paid/unpaid months with details)
    - Group information
    - Parent information

    Filters:
    - from_date: Start date for payment history analysis (defaults to start of current year)
    - to_date: End date for payment history analysis (defaults to today)
    - group_id: Filter students by specific group
    - status: Filter by student status (e.g., 'active', 'archived')
    """
    from app.models.domain import Contract, Group, Parent
    from app.models.enums import ContractStatus, PaymentStatus
    from datetime import date as date_type
    from dateutil.relativedelta import relativedelta
    from sqlalchemy.orm import selectinload

    # Default date range
    today = date_type.today()
    if from_date is None:
        from_date = date_type(today.year, 1, 1)  # Start of current year
    if to_date is None:
        to_date = today

    if from_date > to_date:
        raise HTTPException(status_code=400, detail="from_date must be before or equal to to_date")

    # Calculate all months in the date range
    all_months = []
    current_date = from_date.replace(day=1)
    end_date = to_date.replace(day=1)

    while current_date <= end_date:
        all_months.append((current_date.year, current_date.month))
        current_date += relativedelta(months=1)

    # Build student query with filters
    students_query = select(Student).options(
        selectinload(Student.group),
        selectinload(Student.parents),
        selectinload(Student.contracts)
    )

    if group_id:
        students_query = students_query.where(Student.group_id == group_id)

    if status:
        students_query = students_query.where(Student.status == status)

    students_result = await db.execute(students_query)
    students = students_result.scalars().all()

    # Prepare data for Excel
    student_data_list = []

    for student in students:
        # Get all contracts for this student
        contracts = student.contracts

        # Get parent information
        parent_names = ", ".join([f"{p.first_name} {p.last_name}" for p in student.parents]) if student.parents else "N/A"
        parent_phones = ", ".join([p.phone for p in student.parents]) if student.parents else "N/A"

        # Get group name
        group_name = student.group.name if student.group else "N/A"

        # Process each contract
        if not contracts:
            # Student without contracts
            student_data_list.append({
                "student_id": student.id,
                "first_name": student.first_name,
                "last_name": student.last_name,
                "date_of_birth": student.date_of_birth.strftime("%Y-%m-%d"),
                "phone": student.phone or "N/A",
                "address": student.address or "N/A",
                "status": student.status.value,
                "group": group_name,
                "parent_names": parent_names,
                "parent_phones": parent_phones,
                "contract_number": "N/A",
                "contract_start": "N/A",
                "contract_end": "N/A",
                "contract_status": "N/A",
                "monthly_fee": 0,
                "terminated_at": "N/A",
                "termination_reason": "N/A",
                "paid_months": "N/A",
                "unpaid_months": "N/A",
                "total_expected": 0,
                "total_paid": 0,
                "debt_amount": 0,
            })
        else:
            for contract in contracts:
                # Calculate payment status for this contract
                total_expected = 0
                total_paid = 0
                paid_months_list = []
                unpaid_months_list = []

                for year_val, month_val in all_months:
                    target_date = date_type(year_val, month_val, 1)

                    # Check if this month falls within the contract period
                    effective_end_date = contract.end_date
                    if contract.terminated_at:
                        termination_date = contract.terminated_at.date()
                        if termination_date < effective_end_date:
                            effective_end_date = termination_date

                    contract_start_month = contract.start_date.replace(day=1)
                    contract_end_month = effective_end_date.replace(day=1)

                    if contract_start_month <= target_date <= contract_end_month:
                        # Month is within contract period
                        total_expected += float(contract.monthly_fee)

                        # Check if student has paid for this month
                        # Cast payment_months to JSONB to avoid type mismatch
                        payment_result = await db.execute(
                            select(func.sum(Transaction.amount)).where(
                                Transaction.student_id == student.id,
                                Transaction.contract_id == contract.id,
                                Transaction.status == PaymentStatus.SUCCESS,
                                Transaction.payment_year == year_val,
                                cast(Transaction.payment_months, JSONB).op('@>')(cast([month_val], JSONB))
                            )
                        )
                        month_paid = payment_result.scalar() or 0

                        month_str = f"{year_val}-{month_val:02d}"
                        if month_paid >= contract.monthly_fee:
                            paid_months_list.append(month_str)
                            total_paid += float(month_paid)
                        else:
                            unpaid_months_list.append(month_str)
                            total_paid += float(month_paid)

                debt_amount = max(total_expected - total_paid, 0)

                # Format termination info
                terminated_at_str = contract.terminated_at.strftime("%Y-%m-%d") if contract.terminated_at else "N/A"
                termination_reason = contract.termination_reason if contract.termination_reason else "N/A"

                student_data_list.append({
                    "student_id": student.id,
                    "first_name": student.first_name,
                    "last_name": student.last_name,
                    "date_of_birth": student.date_of_birth.strftime("%Y-%m-%d"),
                    "phone": student.phone or "N/A",
                    "address": student.address or "N/A",
                    "status": student.status.value,
                    "group": group_name,
                    "parent_names": parent_names,
                    "parent_phones": parent_phones,
                    "contract_number": contract.contract_number or "N/A",
                    "contract_start": contract.start_date.strftime("%Y-%m-%d"),
                    "contract_end": contract.end_date.strftime("%Y-%m-%d"),
                    "contract_status": contract.status.value,
                    "monthly_fee": float(contract.monthly_fee),
                    "terminated_at": terminated_at_str,
                    "termination_reason": termination_reason,
                    "paid_months": ", ".join(paid_months_list) if paid_months_list else "None",
                    "unpaid_months": ", ".join(unpaid_months_list) if unpaid_months_list else "None",
                    "total_expected": total_expected,
                    "total_paid": total_paid,
                    "debt_amount": debt_amount,
                })

    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Student Data"

    # Define header style
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Define headers
    headers = [
        "Student ID",
        "First Name",
        "Last Name",
        "Date of Birth",
        "Phone",
        "Address",
        "Status",
        "Group",
        "Parent Names",
        "Parent Phones",
        "Contract Number",
        "Contract Start",
        "Contract End",
        "Contract Status",
        "Monthly Fee",
        "Terminated At",
        "Termination Reason",
        "Paid Months",
        "Unpaid Months",
        "Total Expected",
        "Total Paid",
        "Debt Amount",
    ]

    # Write headers
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    # Write data
    for row_num, student_data in enumerate(student_data_list, 2):
        ws.cell(row=row_num, column=1, value=student_data["student_id"])
        ws.cell(row=row_num, column=2, value=student_data["first_name"])
        ws.cell(row=row_num, column=3, value=student_data["last_name"])
        ws.cell(row=row_num, column=4, value=student_data["date_of_birth"])
        ws.cell(row=row_num, column=5, value=student_data["phone"])
        ws.cell(row=row_num, column=6, value=student_data["address"])
        ws.cell(row=row_num, column=7, value=student_data["status"])
        ws.cell(row=row_num, column=8, value=student_data["group"])
        ws.cell(row=row_num, column=9, value=student_data["parent_names"])
        ws.cell(row=row_num, column=10, value=student_data["parent_phones"])
        ws.cell(row=row_num, column=11, value=student_data["contract_number"])
        ws.cell(row=row_num, column=12, value=student_data["contract_start"])
        ws.cell(row=row_num, column=13, value=student_data["contract_end"])
        ws.cell(row=row_num, column=14, value=student_data["contract_status"])
        ws.cell(row=row_num, column=15, value=student_data["monthly_fee"])
        ws.cell(row=row_num, column=16, value=student_data["terminated_at"])
        ws.cell(row=row_num, column=17, value=student_data["termination_reason"])
        ws.cell(row=row_num, column=18, value=student_data["paid_months"])
        ws.cell(row=row_num, column=19, value=student_data["unpaid_months"])
        ws.cell(row=row_num, column=20, value=student_data["total_expected"])
        ws.cell(row=row_num, column=21, value=student_data["total_paid"])
        ws.cell(row=row_num, column=22, value=student_data["debt_amount"])

    # Adjust column widths
    column_widths = {
        'A': 12, 'B': 15, 'C': 15, 'D': 15, 'E': 15, 'F': 30,
        'G': 12, 'H': 20, 'I': 25, 'J': 20, 'K': 18, 'L': 15,
        'M': 15, 'N': 15, 'O': 12, 'P': 15, 'Q': 20, 'R': 50,
        'S': 50, 'T': 15, 'U': 15, 'V': 15
    }

    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # Add summary row
    if student_data_list:
        summary_row = len(student_data_list) + 3
        ws.cell(row=summary_row, column=1, value="TOTALS").font = Font(bold=True)
        ws.cell(row=summary_row, column=20, value=sum(d["total_expected"] for d in student_data_list)).font = Font(bold=True)
        ws.cell(row=summary_row, column=21, value=sum(d["total_paid"] for d in student_data_list)).font = Font(bold=True)
        ws.cell(row=summary_row, column=22, value=sum(d["debt_amount"] for d in student_data_list)).font = Font(bold=True)

        # Add metadata
        metadata_row = summary_row + 2
        ws.cell(row=metadata_row, column=1, value=f"Report Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}").font = Font(italic=True)
        ws.cell(row=metadata_row + 1, column=1, value=f"Period: {from_date.strftime('%Y-%m-%d')} to {to_date.strftime('%Y-%m-%d')}").font = Font(italic=True)
        ws.cell(row=metadata_row + 2, column=1, value=f"Total Students: {len(set(d['student_id'] for d in student_data_list))}").font = Font(italic=True)
        ws.cell(row=metadata_row + 3, column=1, value=f"Total Contracts: {len(student_data_list)}").font = Font(italic=True)

    # Save to BytesIO
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)

    # Generate filename
    group_suffix = f"_group{group_id}" if group_id else ""
    status_suffix = f"_{status}" if status else ""
    date_str = f"{from_date.strftime('%Y%m%d')}_{to_date.strftime('%Y%m%d')}"
    filename = f"comprehensive_student_data_{date_str}{group_suffix}{status_suffix}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    # Return as streaming response
    return StreamingResponse(
        excel_file,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.post("", response_model=DataResponse[StudentRead], dependencies=[Depends(require_permission(PERM_STUDENTS_EDIT))])
async def create_student(
    data: StudentCreate,
    db: Annotated[AsyncSession, Depends(get_db)],
):
    if data.face_id:
        existing_face_id = await db.execute(select(Student).where(Student.face_id == data.face_id))
        if existing_face_id.scalar_one_or_none():
            raise HTTPException(status_code=400, detail="Face ID already exists. Please use a unique Face ID")

    if data.group_id:
        from app.models.domain import Group
        group_result = await db.execute(select(Group).where(Group.id == data.group_id))
        group = group_result.scalar_one_or_none()
        if not group:
            raise HTTPException(status_code=404, detail=f"Group with ID {data.group_id} not found")

        # Check group capacity - only count ACTIVE students
        if data.status == StudentStatus.ACTIVE:
            active_students_count = await db.execute(
                select(func.count(Student.id)).where(
                    Student.group_id == data.group_id,
                    Student.status == StudentStatus.ACTIVE
                )
            )
            current_count = active_students_count.scalar() or 0

            if current_count >= group.capacity:
                raise HTTPException(
                    status_code=409,
                    detail=f"Group '{group.name}' is at full capacity ({group.capacity} students). "
                           f"Cannot add more active students. Consider adding to waiting list instead."
                )

    student = Student(**data.model_dump())
    db.add(student)
    await db.commit()
    await db.refresh(student)
    return DataResponse(data=StudentRead.model_validate(student))

from dateutil.relativedelta import relativedelta
from decimal import Decimal, InvalidOperation
from fastapi import HTTPException
import asyncio

@router.post("/create-with-contract")
async def create_student_with_contract(
    user: Annotated[User, Depends(require_permission(PERM_STUDENTS_EDIT))],
    db: Annotated[AsyncSession, Depends(get_db)],

    # ========== JSON DATA ==========
    student_data: str = Form(..., description="Student data as JSON"),
    contract_data: str = Form(..., description="Contract data as JSON"),

    # ========== DOCUMENT FILES (FormData) ==========
    passport_copy: UploadFile = File(..., description="Profile photo / Father passport (used as profile image)"),
    form_086: UploadFile = File(..., description="Medical form 086 file"),
    heart_checkup: UploadFile = File(..., description="Heart checkup document file"),
    birth_certificate: UploadFile = File(..., description="Birth certificate (front side)"),
    contract_image_1: UploadFile | None = File(None, description="Birth certificate back (optional)"),
    contract_image_2: UploadFile = File(..., description="Father passport front (mandatory)"),
    contract_image_3: UploadFile | None = File(None, description="Father passport back (optional)"),
    contract_image_4: UploadFile = File(..., description="Mother passport front (mandatory)"),
    contract_image_5: UploadFile | None = File(None, description="Mother passport back (optional)"),

):
    """
    Create student with contract and all documents in ONE operation.

    **Contract number must be provided by admin**
    - Use GET /contracts/next-available/{group_id} to get the next number
    - Admin must enter the contract_number manually
    - Numbers must be sequential (N1, N2, N3, etc.)
    - Once used, numbers cannot be reused (even if terminated)

    **student_data JSON structure:**
    ```json
    {
        "first_name": "Alvaro",
        "last_name": "Marata",
        "date_of_birth": "2010-12-06",
        "phone": "998901234567",
        "address": "Toshkent shahar",
        "status": "active",
        "group_id": 1
    }
    ```

    **contract_data JSON structure:**
    ```json
     {
          "contract_number": "5-2014B1",
          "student": {
            "student_image": "student_photo.png",
            "student_fio": "Юсупов Абдулборий Баҳодирович",
            "birth_year": "2012",
            "student_address": "Тошкент ш. Чилонзор т. Лутфий кўчаси 61-уй",
            "dad_occupation": "Тадбиркор",
            "mom_occupation": "Уй бекаси",
            "dad_phone_number": "(33) 135-80-09",
            "mom_phone_number": "(78) 162-16-14",
            "mom_fullname": "Бахриддинова Гулова Баҳромовна"
          },
          "sana": {
            "kun": "06",
            "oy": "Декабр",
            "yil": "2025"
          },
          "buyurtmachi": {
            "fio": "Юсупов Абдулборий Баҳодирович",
            "pasport_seriya": "AA 1234567",
            "pasport_kim_bergan": "Тошкент ш. Чилонзор т. ИИББ бўлими",
            "pasport_qachon_bergan": "15.03.2018",
            "manzil": "Тошкент ш., Чилонзор тумани, Лутфий кўчаси 61-уй",
            "telefon": "+998 (33) 135-80-09"
          },
          "tarbiyalanuvchi": {
            "fio": "Юсупов Абдулборий Баҳодирович",
            "tugilganlik_guvohnoma": "I-AA 9876543",
            "tugilganlik_yil": 2011,
            "guvohnoma_kim_bergan": "Тошкент ш. ФҲБ Чилонзор т. бўлими",
            "guvohnoma_qachon_bergan": "12.04.2012"
          },
          "shartnoma_muddati": {
            "boshlanish": "2026-01-06",
            "tugash": "2026-12-06",
            "yil": "2025"
          },
          "tolov": {
            "oylik_narx": "600 000",
            "oylik_narx_sozlar": "олти юз минг"
          }
    }

    ```

    Complete workflow:
    1. Parse JSON data from student_data and contract_data
    2. Upload 9 files to AWS S3 automatically
    3. Create student and contract with ACTIVE status
    4. Generate PDF contract using contractdoc.py logic
    5. Return generated PDF file
    """
    from app.models.domain import Group, Contract, WaitingList
    from app.models.enums import ContractStatus, StudentStatus
    from app.services.contract_allocation import (
        is_group_full,
        ContractNumberAllocationError,
        validate_contract_number
    )
    import json
    import os
    import tempfile
    from datetime import datetime

    # Parse JSON data
    try:
        student_info = json.loads(student_data)
        contract_info = json.loads(contract_data)
    except json.JSONDecodeError as e:
        raise HTTPException(status_code=400, detail=f"Invalid JSON: {str(e)}")

    # Extract required fields from student_data
    first_name = student_info.get("first_name")
    last_name = student_info.get("last_name")
    date_of_birth = student_info.get("date_of_birth")
    phone = student_info.get("phone")
    address = student_info.get("address")
    status = student_info.get("status", "active")
    group_id = student_info.get("group_id")

    # Validate required fields
    if not all([first_name, last_name, date_of_birth, group_id]):
        raise HTTPException(
            status_code=400,
            detail="Missing required fields in student_data: first_name, last_name, date_of_birth, group_id"
        )
    contract_images_urls = []
    if contract_image_1:
        contract_images_urls.append(await upload_image_to_s3(contract_image_1, "contracts"))
    if contract_image_2:
        contract_images_urls.append(await upload_image_to_s3(contract_image_2, "contracts"))
    if contract_image_3:
        contract_images_urls.append(await upload_image_to_s3(contract_image_3, "contracts"))
    if contract_image_4:
        contract_images_urls.append(await upload_image_to_s3(contract_image_4, "contracts"))
    if contract_image_5:
        contract_images_urls.append(await upload_image_to_s3(contract_image_5, "contracts"))

    # Extract contract fields
    buyurtmachi = contract_info.get("buyurtmachi", {})
    studentInfo=contract_info.get("student", {})
    tarbiyalanuvchi = contract_info.get("tarbiyalanuvchi", {})
    shartnoma_muddati = contract_info.get("shartnoma_muddati", {})
    tolov = contract_info.get("tolov", {})

    # Extract contract_number (required, admin must enter it manually)
    contract_number = contract_info.get("contract_number")
    if not contract_number:
        raise HTTPException(
            status_code=400,
            detail="contract_number is required in contract_data. Use GET /contracts/next-available/{group_id} to get the next available number."
        )

    # Validate group exists and get birth_year from group
    group_result = await db.execute(select(Group).where(Group.id == group_id))
    group = group_result.scalar_one_or_none()

    if not group:
        raise HTTPException(status_code=404, detail=f"Group with ID {group_id} not found")

    # Use group's birth_year (not from user input)
    birth_year = group.birth_year

    # Check if group is full
    group_full = await is_group_full(db, group_id, birth_year)

    if group_full:
        raise HTTPException(
            status_code=409,
            detail=f"Group '{group.name}' is full (capacity: {group.capacity}). Cannot create contract. Add to waiting list instead."
        )

    # Upload all files to S3
    # Upload all files to S3
    try:
        # Fayl pointerlarini qayta boshiga olish
        for f in [passport_copy, form_086, heart_checkup, birth_certificate, contract_image_1, contract_image_2, contract_image_3, contract_image_4, contract_image_5]:
            if f is not None:
                f.file.seek(0)

        # Asosiy hujjatlar
        passport_copy_url = await upload_image_to_s3(passport_copy, "student-documents")  # Profile image
        form_086_url = await upload_image_to_s3(form_086, "student-documents")
        heart_checkup_url = await upload_image_to_s3(heart_checkup, "student-documents")
        birth_certificate_url = await upload_image_to_s3(birth_certificate, "student-documents")

        # Contract images (passports)
        contract_images_urls = []
        # Index 0: birth certificate back (optional)
        if contract_image_1:
            contract_images_urls.append(await upload_image_to_s3(contract_image_1, "contracts"))
        else:
            contract_images_urls.append(None)

        # Index 1: father passport front (mandatory)
        contract_images_urls.append(await upload_image_to_s3(contract_image_2, "contracts"))

        # Index 2: father passport back (optional)
        if contract_image_3:
            contract_images_urls.append(await upload_image_to_s3(contract_image_3, "contracts"))
        else:
            contract_images_urls.append(None)

        # Index 3: mother passport front (mandatory)
        contract_images_urls.append(await upload_image_to_s3(contract_image_4, "contracts"))

        # Index 4: mother passport back (optional)
        if contract_image_5:
            contract_images_urls.append(await upload_image_to_s3(contract_image_5, "contracts"))
        else:
            contract_images_urls.append(None)


    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error uploading files to S3: {str(e)}")

    # Parse date_of_birth if string
    if isinstance(date_of_birth, str):
        from datetime import datetime
        date_of_birth = datetime.strptime(date_of_birth, "%Y-%m-%d").date()

    # Get current year for archive
    current_year = datetime.now().year

    # Create student
    student_status = StudentStatus.ACTIVE if status.lower() == "active" else StudentStatus.INACTIVE
    student = Student(
        first_name=first_name,
        last_name=last_name,
        date_of_birth=date_of_birth,
        phone=phone,
        address=address,
        status=student_status,
        group_id=group_id,
        archive_year=current_year  # Set current year as archive year
    )
    db.add(student)
    await db.commit()
    await db.refresh(student)

    # Validate the contract number provided by admin
    is_valid, message, sequence_number = await validate_contract_number(
        db, contract_number, group_id, birth_year, current_year
    )

    if not is_valid:
        # Rollback student creation if contract number is invalid
        await db.rollback()
        await db.delete(student)
        await db.commit()
        raise HTTPException(
            status_code=400,
            detail=f"Invalid contract number: {message}. Use GET /contracts/next-available/{group_id} to get the next available number."
        )

    # Parse contract dates
    start_date_str = shartnoma_muddati.get("boshlanish")
    end_date_str = shartnoma_muddati.get("tugash")
    year_val = shartnoma_muddati.get("yil")
    monthly_fee_raw = tolov.get("oylik_narx", 0)


    try:
        if isinstance(monthly_fee_raw, str):
            # Masalan: "600 000" yoki "600,000" yoki "600000"
            cleaned = monthly_fee_raw.replace(" ", "").replace(",", "")
            monthly_fee = Decimal(cleaned)
        else:
            monthly_fee = Decimal(monthly_fee_raw)
    except (InvalidOperation, TypeError, ValueError):
        raise HTTPException(
            status_code=400,
            detail=f"To‘lov miqdati noto‘g‘ri formatda yuborilgan: {monthly_fee_raw!r}"
        )

    # --- Sana formatlarini tekshirish (boshlanish) ---
    try:
        if start_date_str:
            # Foydalanuvchi "2025-01-01" formatda yuboradi
            start_date = datetime.strptime(start_date_str, "%Y-%m-%d").date()
        else:
            raise ValueError("boshlanish sanasi kiritilmagan")
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Sana formati noto‘g‘ri (boshlanish): {str(e)}")

    # --- Tugash sanasi ---
    try:
        if end_date_str and end_date_str.isdigit():
            # Masalan: "31" → 31 dekabr {yil}
            end_date = datetime(int(year_val or start_date.year), 12, int(end_date_str)).date()
        elif end_date_str:
            # Agar foydalanuvchi ISO formatda yuborsa
            end_date = datetime.strptime(end_date_str, "%Y-%m-%d").date()
        else:
            # Tugash kiritilmagan bo‘lsa — 1 yilga uzaytiramiz
            end_date = start_date + relativedelta(years=1)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Sana formati noto‘g‘ri (tugash): {str(e)}")

    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Sana formati noto‘g‘ri (tugash): {str(e)}")

    # Convert data to JSON strings for storage
    contract_images_json_str = json.dumps(contract_images_urls)
    custom_fields_json_str = json.dumps(contract_info, ensure_ascii=False, default=str)
    existing_contract = await db.execute(
        select(Contract).where(Contract.contract_number == contract_number)
    )
    if existing_contract.scalar():
        raise HTTPException(
            status_code=400,
            detail=f"Shartnoma raqami '{contract_number}' allaqachon mavjud."
        )
    # Create contract in ACTIVE status (no signature needed)
    contract = Contract(
        contract_number=contract_number,
        birth_year=birth_year,
        sequence_number=sequence_number,
        start_date=start_date,
        end_date=end_date,
        monthly_fee=monthly_fee,
        status=ContractStatus.ACTIVE,
        student_id=student.id,
        group_id=group_id,
        archive_year=current_year,  # Set current year as archive year
        passport_copy_url=passport_copy_url,
        form_086_url=form_086_url,
        heart_checkup_url=heart_checkup_url,
        birth_certificate_url=birth_certificate_url,
        contract_images_urls=contract_images_json_str,
        custom_fields=custom_fields_json_str
    )

    db.add(contract)
    await db.commit()
    await db.refresh(contract)

    # Prepare data for PDF generation (contractdoc.py format)
    # Parse sana from start_date
    sana_obj = start_date
    months_uz = {
        1: "январь", 2: "февраль", 3: "март", 4: "апрель",
        5: "май", 6: "июнь", 7: "июль", 8: "август",
        9: "сентябрь", 10: "октябрь", 11: "ноябрь", 12: "декабрь"
    }

    pdf_data = {
        "shartnoma_raqami": contract_number,
        "student": contract_info.get("student", {}),
        "sana": {
            "kun": f"{sana_obj.day:02d}",
            "oy": months_uz.get(sana_obj.month, ""),
            "yil": str(sana_obj.year)
        },
        "buyurtmachi": buyurtmachi,
        "tarbiyalanuvchi": tarbiyalanuvchi,
        "shartnoma_muddati": {
            "boshlanish": start_date.strftime('«%d» %B'),
            "tugash": end_date.strftime('«%d» %B'),
            "yil": str(start_date.year)
        },
        "tolov": {
            "oylik_narx": f"{monthly_fee:,.0f}".replace(",", " "),
            "oylik_narx_sozlar": "sum"  # You can add number-to-words conversion here
        }
    }
    # passport_copy ni profile image sifatida ishlatamiz
    pdf_data["student"]["student_image"] = passport_copy_url

    # Generate PDF with attachments (images at the end)
    import time

    temp_pdf = tempfile.NamedTemporaryFile(delete=False, suffix=".pdf")
    pdf_path = temp_pdf.name
    temp_pdf.close()

    try:
        # Add URLs to pdf_data for contract_pdf.py to use
        pdf_data["passport_copy_url"] = passport_copy_url
        pdf_data["form_086_url"] = form_086_url
        pdf_data["heart_checkup_url"] = heart_checkup_url
        pdf_data["birth_certificate_url"] = birth_certificate_url
        pdf_data["contract_images_urls"] = contract_images_urls

        generator = ContractPDFGenerator(pdf_data)

        # Generate PDF with all attachments (contract_pdf.py handles ordering)
        final_pdf_path = generator.generate(pdf_path)

        # Check if generation was successful
        if not final_pdf_path or not isinstance(final_pdf_path, (str, os.PathLike)):
            raise ValueError(f"PDF generation failed: {type(final_pdf_path)}")

        # Upload final PDF to S3
        pdf_s3_url = upload_pdf_to_s3(final_pdf_path, contract_number)

        # Update contract with PDF URL
        contract.final_pdf_url = pdf_s3_url
        await db.commit()

        # Clean up temp files
        try:
            time.sleep(0.2)
            if os.path.exists(pdf_path):
                os.unlink(pdf_path)
            if os.path.exists(final_pdf_path):
                os.unlink(final_pdf_path)
        except:
            pass

        # Return success response with contract and PDF URL
        return DataResponse(data={
            "message": "Student and contract created successfully",
            "student_id": student.id,
            "contract_id": contract.id,
            "contract_number": contract_number,
            "pdf_url": pdf_s3_url
        })

    except Exception as e:
        # Clean up temp files on error
        try:
            time.sleep(0.2)
            if os.path.exists(pdf_path):
                os.unlink(pdf_path)
            if 'final_pdf_path' in locals() and os.path.exists(final_pdf_path):
                os.unlink(final_pdf_path)
        except:
            pass

        raise HTTPException(status_code=500, detail=f"Error generating PDF: {str(e)}")


@router.get("/{student_id}", response_model=DataResponse[StudentRead], dependencies=[Depends(require_permission(PERM_STUDENTS_VIEW))])
async def get_student(
    student_id: int,
    db: Annotated[AsyncSession, Depends(get_db)],
):
    result = await db.execute(select(Student).where(Student.id == student_id))
    student = result.scalar_one_or_none()

    if not student:
        raise HTTPException(status_code=404, detail="Student not found")

    return DataResponse(data=StudentRead.model_validate(student))


@router.get("/fullinfo/{student_id}", response_model=DataResponse[StudentFullInfo], dependencies=[Depends(require_permission(PERM_STUDENTS_VIEW))])
async def get_student_full_info(
    student_id: int,
    db: Annotated[AsyncSession, Depends(get_db)],
):
    """
    Get complete student information including:
    - Student details
    - Parents
    - Contracts
    - Group
    - Coach (teacher)
    - Payment history (transactions)
    - Attendance records
    """
    from app.models.domain import Contract, Parent, Group
    from app.models.auth import User
    from sqlalchemy.orm import selectinload

    # Fetch student
    student_result = await db.execute(select(Student).where(Student.id == student_id))
    student = student_result.scalar_one_or_none()

    if not student:
        raise HTTPException(status_code=404, detail="Student not found")

    # Fetch parents
    parents_result = await db.execute(select(Parent).where(Parent.student_id == student_id))
    parents = parents_result.scalars().all()

    # Fetch contracts
    contracts_result = await db.execute(select(Contract).where(Contract.student_id == student_id))
    contracts = contracts_result.scalars().all()

    # Fetch group and coach if student has a group
    group = None
    coach = None
    if student.group_id:
        group_result = await db.execute(select(Group).where(Group.id == student.group_id))
        group = group_result.scalar_one_or_none()

        if group and group.coach_id:
            coach_result = await db.execute(
                select(User).where(User.id == group.coach_id)
            )
            coach = coach_result.scalar_one_or_none()

    # Fetch transactions
    transactions_result = await db.execute(
        select(Transaction).where(Transaction.student_id == student_id).order_by(Transaction.created_at.desc())
    )
    transactions = transactions_result.scalars().all()

    # Fetch attendance records
    attendances_result = await db.execute(
        select(Attendance).where(Attendance.student_id == student_id).order_by(Attendance.created_at.desc())
    )
    attendances = attendances_result.scalars().all()

    # Build the full info response
    from app.schemas.group import GroupRead
    from app.schemas.auth import UserRead

    full_info = StudentFullInfo(
        student=StudentRead.model_validate(student),
        parents=[ParentRead.model_validate(p) for p in parents],
        contracts=[ContractRead.model_validate(c) for c in contracts],
        group=GroupRead.model_validate(group) if group else None,
        coach=UserRead.model_validate(coach) if coach else None,
        transactions=[TransactionRead.model_validate(t) for t in transactions],
        attendances=[AttendanceRead.model_validate(a) for a in attendances],
    )

    return DataResponse(data=full_info)


@router.patch("/{student_id}", response_model=DataResponse[StudentRead], dependencies=[Depends(require_permission(PERM_STUDENTS_EDIT))])
async def update_student(
    student_id: int,
    data: StudentUpdate,
    db: Annotated[AsyncSession, Depends(get_db)],
):
    result = await db.execute(select(Student).where(Student.id == student_id))
    student = result.scalar_one_or_none()

    if not student:
        raise HTTPException(status_code=404, detail="Student not found")

    update_data = data.model_dump(exclude_unset=True)

    # Prevent changing student's group once assigned
    if "group_id" in update_data and student.group_id is not None:
        if update_data["group_id"] != student.group_id:
            raise HTTPException(
                status_code=400,
                detail="Cannot change student's group. Students cannot be transferred between groups."
            )

    if "face_id" in update_data and update_data["face_id"] is not None:
        existing_face_id = await db.execute(
            select(Student).where(Student.face_id == update_data["face_id"], Student.id != student_id)
        )
        if existing_face_id.scalar_one_or_none():
            raise HTTPException(status_code=400, detail="Face ID already exists. Please use a unique Face ID")

    # Check capacity when assigning to a group or changing status to ACTIVE
    target_group_id = update_data.get("group_id", student.group_id)
    target_status = update_data.get("status", student.status)

    # Determine if this update will make student ACTIVE in a group
    will_be_active_in_group = (
        target_group_id is not None and
        target_status == StudentStatus.ACTIVE and
        (student.group_id != target_group_id or student.status != StudentStatus.ACTIVE)
    )

    if will_be_active_in_group:
        from app.models.domain import Group
        group_result = await db.execute(select(Group).where(Group.id == target_group_id))
        group = group_result.scalar_one_or_none()

        if not group:
            raise HTTPException(status_code=404, detail=f"Group with ID {target_group_id} not found")

        # Count current ACTIVE students in target group (excluding this student)
        active_students_count = await db.execute(
            select(func.count(Student.id)).where(
                Student.group_id == target_group_id,
                Student.status == StudentStatus.ACTIVE,
                Student.id != student_id  # Exclude current student from count
            )
        )
        current_count = active_students_count.scalar() or 0

        if current_count >= group.capacity:
            raise HTTPException(
                status_code=409,
                detail=f"Group '{group.name}' is at full capacity ({group.capacity} students). "
                       f"Cannot add more active students. Consider adding to waiting list instead."
            )
    elif "group_id" in update_data and update_data["group_id"] is not None:
        # Just validate group exists if only changing group
        from app.models.domain import Group
        group_result = await db.execute(select(Group).where(Group.id == update_data["group_id"]))
        if not group_result.scalar_one_or_none():
            raise HTTPException(status_code=404, detail=f"Group with ID {update_data['group_id']} not found")

    for field, value in update_data.items():
        setattr(student, field, value)

    await db.commit()
    await db.refresh(student)
    return DataResponse(data=StudentRead.model_validate(student))


@router.get("/{student_id}/contracts", response_model=DataResponse[list[ContractRead]], dependencies=[Depends(require_permission(PERM_STUDENTS_VIEW))])
async def get_student_contracts(
    student_id: int,
    db: Annotated[AsyncSession, Depends(get_db)],
):
    from app.models.domain import Contract
    result = await db.execute(select(Contract).where(Contract.student_id == student_id))
    contracts = result.scalars().all()
    return DataResponse(data=[ContractRead.model_validate(c) for c in contracts])


@router.get("/{student_id}/transactions", response_model=DataResponse[list[TransactionRead]], dependencies=[Depends(require_permission(PERM_STUDENTS_VIEW))])
async def get_student_transactions(
    student_id: int,
    db: Annotated[AsyncSession, Depends(get_db)],
):
    result = await db.execute(select(Transaction).where(Transaction.student_id == student_id))
    transactions = result.scalars().all()
    return DataResponse(data=[TransactionRead.model_validate(t) for t in transactions])


@router.get("/{student_id}/attendance", response_model=DataResponse[list[AttendanceRead]], dependencies=[Depends(require_permission(PERM_STUDENTS_VIEW))])
async def get_student_attendance(
    student_id: int,
    db: Annotated[AsyncSession, Depends(get_db)],
):
    result = await db.execute(select(Attendance).where(Attendance.student_id == student_id))
    attendance = result.scalars().all()
    return DataResponse(data=[AttendanceRead.model_validate(a) for a in attendance])


@router.get("/{student_id}/gatelogs", response_model=DataResponse[list[GateLogRead]], dependencies=[Depends(require_permission(PERM_STUDENTS_VIEW))])
async def get_student_gatelogs(
    student_id: int,
    db: Annotated[AsyncSession, Depends(get_db)],
):
    result = await db.execute(select(GateLog).where(GateLog.student_id == student_id))
    logs = result.scalars().all()
    return DataResponse(data=[GateLogRead.model_validate(l) for l in logs])


@router.get("/attendances/all", response_model=DataResponse[list[AttendanceRead]], dependencies=[Depends(require_permission(PERM_ATTENDANCE_VIEW))])
async def get_all_attendances(
    db: Annotated[AsyncSession, Depends(get_db)],
    from_date: Optional[date] = Query(None, description="Start date for filtering (YYYY-MM-DD)"),
    to_date: Optional[date] = Query(None, description="End date for filtering (YYYY-MM-DD)"),
    group_id: Optional[int] = Query(None, description="Filter by group ID"),
    student_id: Optional[int] = Query(None, description="Filter by student ID"),
    page: int = Query(1, ge=1, description="Page number"),
    page_size: int = Query(50, ge=1, le=100, description="Items per page"),
):
    """
    Get all student attendances with filters.

    This endpoint allows authorized users to view all attendance records
    with various filters including date range, group, and student.

    Filters:
    - from_date: Filter by session date (start)
    - to_date: Filter by session date (end)
    - group_id: Filter by specific group
    - student_id: Filter by specific student
    - page: Page number for pagination
    - page_size: Number of records per page (max 100)

    Note: This is for viewing marked attendances (coach-created).
    Turnstile/gate attendance is handled separately.
    """
    from sqlalchemy.orm import selectinload

    # Build query for all attendances
    attendance_query = select(Attendance).options(
        selectinload(Attendance.student),
        selectinload(Attendance.session).selectinload(Session.group),
        selectinload(Attendance.marked_by)
    )

    # Apply date filters via session
    if from_date or to_date:
        attendance_query = attendance_query.join(Session)
        if from_date:
            attendance_query = attendance_query.where(Session.session_date >= from_date)
        if to_date:
            attendance_query = attendance_query.where(Session.session_date <= to_date)

    # Apply group filter via session
    if group_id:
        if not (from_date or to_date):  # Only join if not already joined
            attendance_query = attendance_query.join(Session)
        attendance_query = attendance_query.where(Session.group_id == group_id)

    # Apply student filter
    if student_id:
        attendance_query = attendance_query.where(Attendance.student_id == student_id)

    # Order by most recent first
    attendance_query = attendance_query.order_by(Attendance.created_at.desc())

    # Get total count for pagination
    count_query = select(func.count()).select_from(attendance_query.subquery())
    total_result = await db.execute(count_query)
    total = total_result.scalar() or 0

    # Apply pagination
    offset = (page - 1) * page_size
    attendance_query = attendance_query.offset(offset).limit(page_size)

    # Execute query
    attendances_result = await db.execute(attendance_query)
    attendances = attendances_result.scalars().all()

    return DataResponse(
        data=[AttendanceRead.model_validate(a) for a in attendances],
        meta=PaginationMeta(
            page=page,
            page_size=page_size,
            total=total,
            total_pages=(total + page_size - 1) // page_size,
        ),
    )


@router.delete("/{student_id}", response_model=DataResponse[dict], dependencies=[Depends(require_permission(PERM_STUDENTS_EDIT))])
async def delete_student(
    student_id: int,
    db: Annotated[AsyncSession, Depends(get_db)],
):
    """
    Soft delete a student by setting their status to DELETED.
    The student is not actually removed from the database.
    """
    result = await db.execute(select(Student).where(Student.id == student_id))
    student = result.scalar_one_or_none()

    if not student:
        raise HTTPException(status_code=404, detail="Student not found")

    # Soft delete: set status to DELETED instead of actually deleting
    student.status = StudentStatus.DELETED
    await db.commit()

    return DataResponse(data={"message": "Student deleted successfully"})


@router.post("/bulk-delete", response_model=DataResponse[dict], dependencies=[Depends(require_permission(PERM_STUDENTS_EDIT))])
async def bulk_delete_students(
    student_ids: list[int],
    db: Annotated[AsyncSession, Depends(get_db)],
):
    """
    Soft delete multiple students by setting their status to DELETED.
    Students are not actually removed from the database.
    """
    if not student_ids:
        raise HTTPException(status_code=400, detail="No student IDs provided")

    deleted_count = 0
    errors = []

    for student_id in student_ids:
        try:
            result = await db.execute(select(Student).where(Student.id == student_id))
            student = result.scalar_one_or_none()

            if not student:
                errors.append({"student_id": student_id, "error": "Student not found"})
                continue

            # Soft delete: set status to DELETED instead of actually deleting
            student.status = StudentStatus.DELETED
            deleted_count += 1
        except Exception as e:
            errors.append({"student_id": student_id, "error": str(e)})

    await db.commit()

    return DataResponse(data={
        "message": f"Deleted {deleted_count} student(s)",
        "deleted_count": deleted_count,
        "total_requested": len(student_ids),
        "errors": errors if errors else None
    })
