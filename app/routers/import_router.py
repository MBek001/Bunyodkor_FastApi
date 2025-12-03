from typing import Annotated
from datetime import datetime, date
from fastapi import APIRouter, Depends, UploadFile, File, HTTPException
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select
from openpyxl import load_workbook
from io import BytesIO
from app.core.db import get_db
from app.core.permissions import PERM_STUDENTS_EDIT, PERM_FINANCE_TRANSACTIONS_MANUAL
from app.schemas.common import DataResponse
from app.deps import require_permission, CurrentUser
from app.models.domain import Student, Parent, Contract, Group
from app.models.finance import Transaction
from app.models.enums import StudentStatus, ContractStatus, PaymentStatus, PaymentSource

router = APIRouter(prefix="/import", tags=["Import"])


@router.post("/students", response_model=DataResponse[dict], dependencies=[Depends(require_permission(PERM_STUDENTS_EDIT))])
async def import_students(
    file: UploadFile = File(...),
    db: Annotated[AsyncSession, Depends(get_db)] = None,
):
    """
    Import students from Excel file with all related data.

    Expected columns:
    - first_name: Student first name (required)
    - last_name: Student last name (required)
    - date_of_birth: Format YYYY-MM-DD (required)
    - phone: Student phone number
    - address: Student address
    - face_id: Unique face ID
    - status: ACTIVE, INACTIVE, GRADUATED, EXPELLED (default: ACTIVE)
    - group_name: Name of the group to assign student to
    - parent_first_name: Parent first name
    - parent_last_name: Parent last name
    - parent_phone: Parent phone number
    - parent_email: Parent email
    - parent_relationship: Mother, Father, Guardian, etc.
    - contract_number: Unique contract number
    - contract_start_date: Format YYYY-MM-DD
    - contract_end_date: Format YYYY-MM-DD
    - monthly_fee: Monthly fee amount
    """
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Only Excel files (.xlsx, .xls) are supported")

    try:
        # Read Excel file
        contents = await file.read()
        workbook = load_workbook(BytesIO(contents))
        sheet = workbook.active

        # Get header row
        headers = [cell.value for cell in sheet[1]]

        success_count = 0
        error_count = 0
        errors = []

        # Process each row (skip header)
        for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            try:
                # Create dictionary from row
                row_data = dict(zip(headers, row))

                # Skip empty rows
                if not row_data.get('first_name') or not row_data.get('last_name'):
                    continue

                # Parse date_of_birth
                dob_str = row_data.get('date_of_birth')
                if isinstance(dob_str, str):
                    date_of_birth = datetime.strptime(dob_str, '%Y-%m-%d').date()
                elif isinstance(dob_str, datetime):
                    date_of_birth = dob_str.date()
                elif isinstance(dob_str, date):
                    date_of_birth = dob_str
                else:
                    raise ValueError(f"Invalid date_of_birth format in row {row_num}")

                # Check if face_id already exists
                if row_data.get('face_id'):
                    existing_face_id = await db.execute(
                        select(Student).where(Student.face_id == row_data['face_id'])
                    )
                    if existing_face_id.scalar_one_or_none():
                        raise ValueError(f"Face ID {row_data['face_id']} already exists")

                # Get or find group
                group_id = None
                if row_data.get('group_name'):
                    group_result = await db.execute(
                        select(Group).where(Group.name == row_data['group_name'])
                    )
                    group = group_result.scalar_one_or_none()
                    if group:
                        group_id = group.id
                    else:
                        raise ValueError(f"Group '{row_data['group_name']}' not found")

                # Create student
                student_status = row_data.get('status', 'ACTIVE')
                if isinstance(student_status, str):
                    try:
                        student_status = StudentStatus[student_status.upper()]
                    except KeyError:
                        student_status = StudentStatus.ACTIVE

                student = Student(
                    first_name=row_data['first_name'],
                    last_name=row_data['last_name'],
                    date_of_birth=date_of_birth,
                    phone=row_data.get('phone'),
                    address=row_data.get('address'),
                    face_id=row_data.get('face_id'),
                    status=student_status,
                    group_id=group_id,
                )
                db.add(student)
                await db.flush()  # Get student ID

                # Create parent if data provided
                if row_data.get('parent_first_name') and row_data.get('parent_last_name') and row_data.get('parent_phone'):
                    parent = Parent(
                        first_name=row_data['parent_first_name'],
                        last_name=row_data['parent_last_name'],
                        phone=row_data['parent_phone'],
                        email=row_data.get('parent_email'),
                        relationship_type=row_data.get('parent_relationship'),
                        student_id=student.id,
                    )
                    db.add(parent)

                # Create contract if data provided
                if row_data.get('contract_number'):
                    # Check if contract number already exists
                    existing_contract = await db.execute(
                        select(Contract).where(Contract.contract_number == row_data['contract_number'])
                    )
                    if existing_contract.scalar_one_or_none():
                        raise ValueError(f"Contract number {row_data['contract_number']} already exists")

                    # Check if student already has active contract
                    active_contract = await db.execute(
                        select(Contract).where(
                            Contract.student_id == student.id,
                            Contract.status == ContractStatus.ACTIVE
                        )
                    )
                    if active_contract.scalar_one_or_none():
                        raise ValueError(f"Student {student.first_name} {student.last_name} already has an active contract")

                    # Parse contract dates
                    contract_start = row_data.get('contract_start_date')
                    if isinstance(contract_start, str):
                        contract_start = datetime.strptime(contract_start, '%Y-%m-%d').date()
                    elif isinstance(contract_start, datetime):
                        contract_start = contract_start.date()

                    contract_end = row_data.get('contract_end_date')
                    if isinstance(contract_end, str):
                        contract_end = datetime.strptime(contract_end, '%Y-%m-%d').date()
                    elif isinstance(contract_end, datetime):
                        contract_end = contract_end.date()

                    if not contract_start or not contract_end:
                        raise ValueError(f"Contract dates required for contract {row_data['contract_number']}")

                    monthly_fee = float(row_data.get('monthly_fee', 0))
                    if monthly_fee <= 0:
                        raise ValueError(f"Invalid monthly_fee for contract {row_data['contract_number']}")

                    contract = Contract(
                        contract_number=row_data['contract_number'],
                        start_date=contract_start,
                        end_date=contract_end,
                        monthly_fee=monthly_fee,
                        status=ContractStatus.ACTIVE,
                        student_id=student.id,
                    )
                    db.add(contract)

                await db.commit()
                success_count += 1

            except Exception as e:
                await db.rollback()
                error_count += 1
                errors.append({
                    "row": row_num,
                    "error": str(e)
                })

        return DataResponse(
            data={
                "message": "Student import completed",
                "filename": file.filename,
                "success_count": success_count,
                "error_count": error_count,
                "errors": errors if errors else None
            }
        )

    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Error processing Excel file: {str(e)}")


@router.post("/payments", response_model=DataResponse[dict], dependencies=[Depends(require_permission(PERM_FINANCE_TRANSACTIONS_MANUAL))])
async def import_payments(
    file: UploadFile = File(...),
    user: CurrentUser = None,
    db: Annotated[AsyncSession, Depends(get_db)] = None,
):
    """
    Import payments from Excel file.

    Expected columns:
    - contract_number: Contract number (required)
    - amount: Payment amount (required)
    - source: CASH, CARD, BANK_TRANSFER, ONLINE, PAYME, CLICK, UZUM (required)
    - payment_year: Year of payment (required)
    - payment_months: Comma-separated months (e.g., "1,2,3" for Jan, Feb, Mar) (required)
    - paid_at: Payment date in format YYYY-MM-DD HH:MM:SS (optional, defaults to now)
    - comment: Payment comment (optional)
    """
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Only Excel files (.xlsx, .xls) are supported")

    try:
        # Read Excel file
        contents = await file.read()
        workbook = load_workbook(BytesIO(contents))
        sheet = workbook.active

        # Get header row
        headers = [cell.value for cell in sheet[1]]

        success_count = 0
        error_count = 0
        errors = []

        # Process each row (skip header)
        for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            try:
                # Create dictionary from row
                row_data = dict(zip(headers, row))

                # Skip empty rows
                if not row_data.get('contract_number'):
                    continue

                # Find contract
                contract_result = await db.execute(
                    select(Contract).where(Contract.contract_number == row_data['contract_number'])
                )
                contract = contract_result.scalar_one_or_none()

                if not contract:
                    raise ValueError(f"Contract '{row_data['contract_number']}' not found")

                # Parse amount
                amount = float(row_data.get('amount', 0))
                if amount <= 0:
                    raise ValueError(f"Invalid amount: {amount}")

                # Parse payment source
                source_str = row_data.get('source', 'CASH').upper()
                try:
                    source = PaymentSource[source_str]
                except KeyError:
                    raise ValueError(f"Invalid payment source: {source_str}")

                # Parse payment year
                payment_year = int(row_data.get('payment_year'))

                # Parse payment months
                payment_months_str = str(row_data.get('payment_months', ''))
                payment_months = [int(m.strip()) for m in payment_months_str.split(',') if m.strip()]

                # Validate months
                for month in payment_months:
                    if month < 1 or month > 12:
                        raise ValueError(f"Invalid month: {month}. Must be between 1 and 12")

                if not payment_months:
                    raise ValueError("payment_months is required")

                # Parse paid_at
                paid_at = row_data.get('paid_at')
                if paid_at:
                    if isinstance(paid_at, str):
                        paid_at = datetime.strptime(paid_at, '%Y-%m-%d %H:%M:%S')
                    elif not isinstance(paid_at, datetime):
                        paid_at = datetime.utcnow()
                else:
                    paid_at = datetime.utcnow()

                # Create transaction
                transaction = Transaction(
                    amount=amount,
                    source=source,
                    status=PaymentStatus.SUCCESS,
                    student_id=contract.student_id,
                    contract_id=contract.id,
                    payment_year=payment_year,
                    payment_months=payment_months,
                    comment=row_data.get('comment'),
                    paid_at=paid_at,
                    created_by_user_id=user.id,
                )
                db.add(transaction)
                await db.commit()
                success_count += 1

            except Exception as e:
                await db.rollback()
                error_count += 1
                errors.append({
                    "row": row_num,
                    "error": str(e)
                })

        return DataResponse(
            data={
                "message": "Payment import completed",
                "filename": file.filename,
                "success_count": success_count,
                "error_count": error_count,
                "errors": errors if errors else None
            }
        )

    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Error processing Excel file: {str(e)}")
